"""
Parser for 3GPP Work Plan Excel files.
Extracts Release 21 (6G) work items and study items status.
"""
import openpyxl
from typing import Dict, List, Optional
import structlog

logger = structlog.get_logger()


class WorkItemParser:
    """Parse 3GPP Work Plan Excel file to extract Release 21 status"""
    
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.workbook = None
        
    def parse(self) -> Dict:
        """
        Parse the Excel file and extract Release 21 work items.
        
        Returns:
            Dict with structure:
            {
                "total_work_items": int,
                "completed": int,
                "in_progress": int,
                "postponed": int,
                "progress_percentage": float,
                "last_updated": str,
                "work_items_by_group": {
                    "RAN1": {"total": int, "completed": int, "progress": float},
                    ...
                }
            }
        """
        try:
            self.workbook = openpyxl.load_workbook(self.excel_path, read_only=True, data_only=True)
            
            # Try to find the right sheet (common names: "Work Items", "Status Report", etc.)
            sheet = self._find_work_items_sheet()
            if not sheet:
                logger.warning("work_item_parser_no_sheet", path=self.excel_path)
                return self._empty_result()
            
            # Parse the sheet
            work_items = self._parse_sheet(sheet)
            
            # Filter for Release 21
            rel21_items = [wi for wi in work_items if self._is_release_21(wi)]
            
            # Aggregate statistics
            result = self._aggregate_statistics(rel21_items)
            
            logger.info("work_item_parsed", total=result["total_work_items"])
            return result
            
        except Exception as e:
            logger.error("work_item_parser_error", error=str(e))
            return self._empty_result()
        finally:
            if self.workbook:
                self.workbook.close()
    
    def _find_work_items_sheet(self) -> Optional[object]:
        """Find the sheet containing work items data"""
        if not self.workbook:
            return None
        
        # Common sheet names
        possible_names = [
            "Work Items", "Status Report", "WI Status", 
            "Work Plan", "Release 21", "Rel-21"
        ]
        
        for sheet_name in self.workbook.sheetnames:
            if any(name.lower() in sheet_name.lower() for name in possible_names):
                return self.workbook[sheet_name]
        
        # Fallback to first sheet
        return self.workbook.worksheets[0] if self.workbook.worksheets else None
    
    def _parse_sheet(self, sheet) -> List[Dict]:
        """Parse work items from sheet"""
        work_items = []
        headers = []
        
        # Find header row (usually first row or row containing "Work Item", "Status", etc.)
        header_row = None
        for row_idx, row in enumerate(sheet.iter_rows(max_row=10, values_only=True), 1):
            if any(cell and isinstance(cell, str) and 
                   any(keyword in cell.lower() for keyword in ["work item", "wi", "status", "release"])
                   for cell in row):
                headers = [str(cell).strip() if cell else "" for cell in row]
                header_row = row_idx
                break
        
        if not header_row:
            logger.warning("work_item_parser_no_headers")
            return []
        
        # Find column indices
        wi_col = self._find_column_index(headers, ["work item", "wi", "acronym", "name"])
        status_col = self._find_column_index(headers, ["status", "state"])
        release_col = self._find_column_index(headers, ["release", "rel", "target release"])
        wg_col = self._find_column_index(headers, ["wg", "working group", "group", "tsg"])
        
        # Parse data rows
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not any(row):
                continue
            
            work_item = {
                "name": str(row[wi_col]).strip() if wi_col is not None and wi_col < len(row) and row[wi_col] else "",
                "status": str(row[status_col]).strip() if status_col is not None and status_col < len(row) and row[status_col] else "",
                "release": str(row[release_col]).strip() if release_col is not None and release_col < len(row) and row[release_col] else "",
                "working_group": str(row[wg_col]).strip() if wg_col is not None and wg_col < len(row) and row[wg_col] else ""
            }
            
            if work_item["name"]:
                work_items.append(work_item)
        
        return work_items
    
    def _find_column_index(self, headers: List[str], keywords: List[str]) -> Optional[int]:
        """Find column index by matching keywords"""
        for idx, header in enumerate(headers):
            if any(keyword in header.lower() for keyword in keywords):
                return idx
        return None
    
    def _is_release_21(self, work_item: Dict) -> bool:
        """Check if work item is for Release 21"""
        release = work_item.get("release", "").lower()
        return any(marker in release for marker in ["21", "rel-21", "rel21", "release 21"])
    
    def _aggregate_statistics(self, work_items: List[Dict]) -> Dict:
        """Aggregate statistics from work items"""
        from datetime import datetime
        
        total = len(work_items)
        completed = 0
        in_progress = 0
        postponed = 0
        
        # Group by working group
        by_group = {}
        
        for wi in work_items:
            status = wi.get("status", "").lower()
            wg = wi.get("working_group", "Other").upper()
            
            # Classify status
            if any(keyword in status for keyword in ["complete", "approved", "finished"]):
                completed += 1
                status_type = "completed"
            elif any(keyword in status for keyword in ["postpone", "delay", "suspend"]):
                postponed += 1
                status_type = "postponed"
            else:
                in_progress += 1
                status_type = "in_progress"
            
            # Aggregate by group
            if wg not in by_group:
                by_group[wg] = {"total": 0, "completed": 0, "in_progress": 0, "postponed": 0}
            
            by_group[wg]["total"] += 1
            by_group[wg][status_type] += 1
        
        # Calculate progress percentages
        progress_percentage = round((completed / total * 100) if total > 0 else 0, 1)
        
        for group in by_group.values():
            group["progress"] = round((group["completed"] / group["total"] * 100) if group["total"] > 0 else 0, 1)
        
        return {
            "total_work_items": total,
            "completed": completed,
            "in_progress": in_progress,
            "postponed": postponed,
            "progress_percentage": progress_percentage,
            "last_updated": datetime.now().strftime("%Y-%m-%d"),
            "work_items_by_group": by_group
        }
    
    def _empty_result(self) -> Dict:
        """Return empty result structure"""
        from datetime import datetime
        return {
            "total_work_items": 0,
            "completed": 0,
            "in_progress": 0,
            "postponed": 0,
            "progress_percentage": 0,
            "last_updated": datetime.now().strftime("%Y-%m-%d"),
            "work_items_by_group": {}
        }
